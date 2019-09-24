from WindPy import *
import os
from openpyxl import load_workbook

w.start()

path_work = "E:/VSCode/Wind_KeChuangBan/Test"
print(os.getcwd())


wb = load_workbook(filename='./Test/1.xlsx')
row = 1

for n in range(1,40):
    ws = wb['Sheet1']
    sector_name_id_index = 'C' + str(n)
    sector_id = ws[sector_name_id_index].value
    print(sector_id)
    str_data_push = "date=2019-05-30;sectorid=" + sector_id + ";field=wind_code,sec_name"
    print(str_data_push)

    data_class = w.wset("sectorconstituent",str_data_push)
    data = data_class.Data
    try:
        data_length = len(data[0])
    except:
        print(sector_id,' null')
    else:
        print(data_length)

        ws = wb['Sheet2']
    
        for i in range(data_length):
            wind_code_index = 'A' + str(row)
            sec_name_index = 'B' + str(row)
            sector_id_index = 'C' + str(row)
            ws[wind_code_index] = data[0][i]
            ws[sec_name_index] = data[1][i]
            ws[sector_id_index] = sector_id
            print(data[0][i],data[1][i],sector_id)
            row = row + 1
wb.save('./Test/1.xlsx')
w.stop()


